"""
convert.py
----------
Convierte el archivo exportado de Azure DevOps ("WAPSA Team - Epics *.xlsx")
al formato que consume el dashboard (data.xlsx, hoja "Hoja2").
Aplica el mismo formato visual (encabezado azul, anchos, auto-filter) que
el archivo de referencia "Tickets - 23-03-2026 WAPSA.xlsx".

Uso:
    python convert.py                               # usa el archivo fuente mas reciente
    python convert.py "WAPSA Team - Epics (6).xlsx" # especifica otro archivo
"""

import sys
from pathlib import Path

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuracion ──────────────────────────────────────────────────────────────
DEFAULT_SOURCE = "WAPSA Team - Epics (5).xlsx"
OUTPUT_FILE    = "data.xlsx"
OUTPUT_SHEET   = "Hoja2"

# Columnas de Hoja2 (orden exacto)
TARGET_COLS = [
    "ID", "Work Item Type", "Title", "Iteration Path", "Estados", "Tags",
    "Start Date", "Finish Date", "Finish Date Chinalco", "Aplicacion",
    "Ticket Chinalco", "Ticket SyC", "Assigned To", "Description",
    "Tipo Caso", "Tipo Evento", "Tipo Estabilizacion", "Parent", "Epic",
]

# Anchos de columna (medidos del archivo de referencia)
COL_WIDTHS = {
    "A": 4.14,   # ID
    "B": 14.0,   # Work Item Type
    "C": 77.0,   # Title
    "D": 40.57,  # Iteration Path
    "E": 19.57,  # Estados
    "F": 15.43,  # Tags
    "G": 17.71,  # Start Date
    "H": 13.0,   # Finish Date
    "I": 15.71,  # Finish Date Chinalco
    "J": 13.0,   # Aplicacion
    "K": 9.29,   # Ticket Chinalco
    "L": 17.71,  # Ticket SyC
    "M": 47.71,  # Assigned To
    "N": 255.71, # Description
    "O": 13.0,   # Tipo Caso
    "P": 35.43,  # Tipo Evento
    "Q": 25.0,   # Tipo Estabilizacion
    "R": 13.0,   # Parent
    "S": 24.71,  # Epic
}

# Estilos del encabezado (segun archivo de referencia)
HEADER_FILL  = PatternFill("solid", fgColor="FF106EBE")
HEADER_FONT  = Font(bold=True, color="FFFFFFFF", size=9)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)
HEADER_BORDER = Border(right=Side(border_style="medium", color="FFFFFF"))

HEADER_ROW_HEIGHT = 24.0
# ──────────────────────────────────────────────────────────────────────────────


def detect_source() -> Path:
    if len(sys.argv) > 1:
        p = Path(sys.argv[1])
        if not p.exists():
            sys.exit(f"[ERROR] Archivo no encontrado: {p}")
        return p

    candidates = sorted(
        Path(".").glob("WAPSA Team - Epics*.xlsx"),
        key=lambda f: f.stat().st_mtime,
        reverse=True,
    )
    if candidates:
        return candidates[0]

    p = Path(DEFAULT_SOURCE)
    if not p.exists():
        sys.exit(
            f"[ERROR] No se encontro el archivo fuente.\n"
            f"   Asegurate de que exista '{DEFAULT_SOURCE}' en esta carpeta\n"
            f"   o pasalo como argumento:  python convert.py <archivo.xlsx>"
        )
    return p


def normalize_col_name(name: str) -> str:
    """Normaliza nombres de columna para tolerar variaciones de exportacion."""
    replacements = {
        "Aplicacion": "Aplicacion",   # quitar acento para columna interna
        "Aplicación": "Aplicacion",
        "Tipo Estabilización": "Tipo Estabilizacion",
        "Tipo Estabilizacion": "Tipo Estabilizacion",
    }
    name = name.strip()
    return replacements.get(name, name)


def apply_formatting(ws, n_rows: int):
    """Aplica formato visual al worksheet: encabezado azul, anchos, auto-filter."""

    # Altura de fila encabezado
    ws.row_dimensions[1].height = HEADER_ROW_HEIGHT

    # Estilo de encabezados
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER

    # Anchos de columna
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # Auto-filter sobre toda la tabla (A1 hasta ultima columna, ultima fila)
    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{n_rows + 1}"


def main():
    source = detect_source()
    print(f"[OK] Leyendo fuente: {source}")

    xf = pd.ExcelFile(source)
    df_raw = pd.read_excel(xf, sheet_name=xf.sheet_names[0])

    # Normalizar nombres de columnas
    df_raw.columns = [normalize_col_name(c) for c in df_raw.columns]

    print(f"   Filas totales: {len(df_raw)}")

    # Mapa Epic ID -> Titulo
    epics = df_raw[df_raw["Work Item Type"] == "Epic"][["ID", "Title"]].copy()
    epic_map: dict = dict(zip(epics["ID"], epics["Title"]))
    print(f"   Epicas encontradas: {len(epic_map)}")

    # Filtrar solo Tickets
    tickets = df_raw[df_raw["Work Item Type"] == "Ticket"].copy()
    print(f"   Tickets: {len(tickets)}")

    # Columna Epic desde el Parent
    def resolve_epic(parent_id):
        if pd.isna(parent_id):
            return ""
        return epic_map.get(int(parent_id), "")

    tickets["Epic"] = tickets["Parent"].apply(resolve_epic)

    # Agregar columnas faltantes
    for col in TARGET_COLS:
        if col not in tickets.columns:
            print(f"   [WARN] Columna no encontrada en fuente, se agrega vacia: '{col}'")
            tickets[col] = ""

    tickets = tickets[TARGET_COLS].reset_index(drop=True)

    # ── Usar el archivo de referencia como plantilla ──────────────────────────
    import shutil, copy
    from openpyxl.styles.fills import PatternFill as PF

    output = Path(OUTPUT_FILE)
    template = Path("Tickets - 23-03-2026 WAPSA.xlsx")

    # Partir de la plantilla si existe, si no del output anterior
    base = template if template.exists() else (output if output.exists() else None)

    if base:
        shutil.copy2(base, output)
        wb = openpyxl.load_workbook(output)
    else:
        wb = openpyxl.Workbook()

    # Asegurar que la hoja existe y limpiarla
    if OUTPUT_SHEET in wb.sheetnames:
        ws = wb[OUTPUT_SHEET]
        # Borrar todas las filas de datos (dejar solo la fila 1 de encabezado)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None
    else:
        ws = wb.create_sheet(OUTPUT_SHEET)

    # Escribir encabezados (fila 1) con valores correctos y re-aplicar formato
    headers = list(tickets.columns)
    for col_idx, header in enumerate(headers, start=1):
        display = (header
                   .replace("Aplicacion", "Aplicaci\u00f3n")
                   .replace("Tipo Estabilizacion", "Tipo Estabilizaci\u00f3n"))
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.fill      = copy.copy(HEADER_FILL)
        cell.font      = copy.copy(HEADER_FONT)
        cell.alignment = copy.copy(HEADER_ALIGN)
        cell.border    = copy.copy(HEADER_BORDER)

    # Escribir datos desde fila 2
    for row_idx, row in enumerate(tickets.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            if isinstance(value, float) and pd.isna(value):
                value = ""
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Aplicar formato (anchos, auto-filter, alto encabezado)
    apply_formatting(ws, n_rows=len(tickets))

    wb.save(output)

    print(f"\n[DONE] {OUTPUT_FILE} actualizado -- {len(tickets)} tickets en hoja '{OUTPUT_SHEET}'")
    print(f"   Epicas unicas: {tickets['Epic'].nunique()}")


if __name__ == "__main__":
    main()
